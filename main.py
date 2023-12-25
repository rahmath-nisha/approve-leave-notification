import getValuesFromExcel
import get_Holiday_List
import os
from datetime import datetime
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from time import sleep
import logging
import requests
import sys

def setup_logger(log_file,name):
    logger = logging.getLogger(name)
    # Check if handlers are already added to the logger
    if len(logger.handlers) == 0:
        logger.setLevel(logging.DEBUG)
        handler = logging.FileHandler(log_file)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', '%H:%M')
        handler.setFormatter(formatter)
        logger.addHandler(handler)
    return logger
today = datetime.today()

log_file = f'C:/Users/ADMIN/Log/LeaveRegisterLog/LeaveRegister_export log on {today.date()}.log'
if (os.path.exists(log_file)):
    print('log created')
    logger = setup_logger(log_file, "my_logger")
else:
    log_file = open(f'C:/Users/ADMIN/Log/LeaveRegisterLog/LeaveRegister_export log on {today.date()}.log', "x")
    print(log_file.name)
    logger = setup_logger(log_file.name, "my_logger")
print(logger)


def take_screenshot(driver, path="screenshot.png"):
    driver.save_screenshot(path)
def download_leave_register():
    try:
        logger.info("exportFileUsingSelenium started execution")
        # Create a new driver instance and add the webdriver
        s = Service('C:/Users/ADMIN/PycharmProjects/leave_register/chromedriver.exe')
        driver = webdriver.Chrome(service=s)

        # Open HRApp
        driver.get('https://digitalseo.hrapp.co/auth')

        today = datetime.today()
        print(today)
        file_Path = f'C:/Users/ADMIN/Leave Register/Company Holiday List {today.year}'
        if (not(os.path.exists(file_Path))):
            logger.info(f'Company Holiday List {today.year} folder not exist')
            get_Holiday_List.export_excel_file(today.year)

        # Maximizing the chrome window
        driver.maximize_window()

        sleep(8)
        # click on the sign with email button
        click_EmailBtn = driver.find_element(By.CSS_SELECTOR, '#signinWithEmailBtnText')
        click_EmailBtn.click()

        # Identify and interact with the username field
        enterEmailId = driver.find_element(By.CSS_SELECTOR, '#formSigninEmailId').send_keys('jaya@seofeeds.com')  # Adjust based on actual element ID or another selector
        sleep(1)

        # click on the Next button
        click_NextBtn = driver.find_element(By.CSS_SELECTOR, '#email-verification-button')
        click_NextBtn.click()
        sleep(3)

        # Identify and interact with the password field
        enterUserPassword = driver.find_element(By.ID, 'formSigninPassword').send_keys('Priya446')  # Adjust based on actual element ID or another selector
        sleep(2)

        # Identify and click the signin button
        click_SignInBtn = driver.find_element(By.ID, 'email-password-submit-button')
        click_SignInBtn.click()
        sleep(20)

        logger.info(f'logged in to HR app')
        # Identify and click the employee button
        click_EmployeeBtn = driver.find_element(By.CLASS_NAME, 'hr-employees')
        click_EmployeeBtn.click()
        sleep(1)

        # inside the employee button we have a option called leave click the click button
        click_EmployeeLeaveBtn = driver.find_element(By.CLASS_NAME, 'hr-employees-leaves')
        click_EmployeeLeaveBtn.click()
        sleep(10)

        # Splitting the date to get day, month, and year
        holiday_book = load_workbook(f'{file_Path}/Location(s) Holidays.xlsx')  # get the activity tracker excel sheet
        holiday_sheet = holiday_book.active  # set the activity_tracker_book sheet as active
        logger.info(f'Iterating the holiday list')
        for cell in holiday_sheet[holiday_sheet.cell(row=1, column=4).column_letter]:
            if(cell.value == today.date()):
                logger.info(f'Leave Register execution stop due to holiday')
                webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
                message_content = f"Leave Register execution stop due to holiday"
                # send the respective employee details to the manager to approve the leave request
                data = {
                    "content": message_content
                }
                response = requests.post(webhook_url, data=data)
                # Check whether the response send or not
                if response.status_code == 204:
                    print("Message sent successfully!")
                else:
                    print(f"Failed to send message. Status code: {response.status_code} {response.text}")
                sys.exit()
        today_Date = today.day
        print(today_Date)
        today_week = today.isoweekday()
        print(today_week)
        global selected_Date
        if (today_week == 7):
            logger.info(f'Execution stop due to is sunday')
            webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
            message_content = f"Leave Register execution stop due to sunday"
            # send the respective employee details to the manager to approve the leave request
            data = {
                "content": message_content
            }
            response = requests.post(webhook_url, data=data)
            # Check whether the response send or not
            if response.status_code == 204:
                print("Message sent successfully!")
            else:
                print(f"Failed to send message. Status code: {response.status_code} {response.text}")
            driver.quit()
            sys.exit()

        # exportLeave button
        click_ExportLeaveBtn = driver.find_element(By.ID, 'exportLeave')
        click_ExportLeaveBtn.click()
        sleep(80)
        screenshot_path = f"C:/Users/ADMIN/ErrorScreenshots/Error_{today.strftime('%Y-%m-%d_%H-%M-%S')}.png"
        take_screenshot(driver, screenshot_path)
        logger.info(f'Screenshot taken and saved to {screenshot_path}')
        sleep(10)
        logger.info(f'Leave register file downloaded')
        # quit the program
        driver.quit()

        logger.info(f'Called the getValuesFromExcel.check_folder_Exist_or_not() method')
        # call the function to perform folder creation, accessing excel file and send message to dicord channel for Leave Register
        getValuesFromExcel.check_folder_Exist_or_not()

    except Exception as e:
        logger.critical(f'Error occurs in the exportFileUsingSelenium {e}')
        webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
        '''# Take a screenshot before sending the error message
        screenshot_path = f"C:/Users/ADMIN/ErrorScreenshots/Error_{today.strftime('%Y-%m-%d_%H-%M-%S')}.png"
        take_screenshot(driver, screenshot_path)
        logger.info(f'Screenshot taken and saved to {screenshot_path}')'''
        print(f"some error occur in the exportFileUsingSelenium code {e}")
        message_content = f"some error occur in the exportFileUsingSelenium code {e}"
        # send the respective employee details to the manager to approve the leave request
        data = {
            "content": message_content
        }
        response = requests.post(webhook_url, data=data)
        # Check whether the response send or not
        if response.status_code == 204:
            print("Message sent successfully!")
        else:
            print(f"Failed to send message. Status code: {response.status_code} {response.text}")

    finally:
        logger.info(f'program executed successfully')
        webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
        print(f"Program executed sucessfully")
        message_content = f"Program executed sucessfully"
        # send the respective employee details to the manager to approve the leave request
        data = {
            "content": message_content
        }
        response = requests.post(webhook_url, data=data)
        # Check whether the response send or not
        if response.status_code == 204:
            print("Message sent successfully!")
        else:
            print(f"Failed to send message. Status code: {response.status_code} {response.text}")

def delete_leave_register_files(folder_path):
    # List all files in the specified folder
    all_files = os.listdir(folder_path)

    # Filter files that contain "Leave Register" in their name
    leave_register_files = [file for file in all_files if "Leave Register" in file]

    # Delete each leave register file
    for file in leave_register_files:
        file_path = os.path.join(folder_path, file)
        os.remove(file_path)
        print(f"Deleted: {file_path}")

# Example usage
download_folder_path = 'C:/Users/ADMIN/Downloads'
delete_leave_register_files(download_folder_path)
download_leave_register()