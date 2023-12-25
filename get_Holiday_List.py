import os
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from time import sleep
import shutil
from openpyxl import load_workbook
import requests
import logging

def setup_logger(log_file,name):
    logger = logging.getLogger(name)
    # Check if handlers are already added to the logger
    if len(logger.handlers) == 0:
        logger.setLevel(logging.DEBUG)
        handler = logging.FileHandler(log_file)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', '%H:%M')
        handler.setFormatter(formatter)
        logger.addHandler(handler)
    return logger
today = datetime.today()

log_file = f'C:/Users/ADMIN/Log/LeaveRegisterLog/LeaveRegister_Holiday log on {today.date()}.log'
if (os.path.exists(log_file)):
    print('log created')
    logger = setup_logger(log_file, "my_logger")
else:
    log_file = open(f"C:/Users/ADMIN/Log/LeaveRegisterLog/LeaveRegister_Holiday log on {today.date()}.log", "x")
    print(log_file)
    logger = setup_logger(log_file.name, "my_logger")
    print(logger)


logger.info("Holiday file does not exist So redirected to holiday get_Holiday_List.export_excel_file(today.year) method")


def remove_duplicate(file_path):
    try:
        logger.info(f"remove_duplicate method gets executed")
        holiday_book = load_workbook(file_path)  # get the activity tracker excel sheet
        activity_tracker_sheet = holiday_book.active  # set the activity_tracker_book sheet as active
        for cell in activity_tracker_sheet[activity_tracker_sheet.cell(row=1, column=1).column_letter]:
            logger.info(f"holiday list iterated")
            if(cell.value == 'Coimbatore' or cell.value == 'Chennai'):
                logger.info(f"removing the duplicates")
                activity_tracker_sheet.delete_rows(cell.row)
                holiday_book.save(file_path)
    except Exception as e:
        logger.critical(f'exception occurs in the remove_duplicate() method {e}')
        webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
        print(f'exception occurs in get_Holiday_List.remove_duplicate() \n {e}')
        message_content = f'exception occurs in get_Holiday_List.remove_duplicate() \n{e}'
        # send the respective employee details to the manager to approve the leave request
        data = {
            "content": message_content
        }
        response = requests.post(webhook_url, data=data)
        # Check whether the response send or not
        if response.status_code == 204:
            print("Message sent successfully!")
        else:
            print(f"Failed to send message. Status code: {response.status_code} {response.text}")

# Move the Activity Tracker file from one location to another
def move_activity_file(current_year):
    try:
        logger.info(f"move_activity_file method executed")
        basePath = f'C:/Users/ADMIN/Leave Register/Company Holiday List {current_year}'         # store the base path in the variable
        file_Path = f'C:/Users/ADMIN/Downloads/Location(s) Holidays.xlsx'            # get the downloaded file path and store it
        if (os.path.exists(file_Path)):                      # check whether the file exist in the downloads folder or not
            logger.info(f"Location(s) Holiday file exist in the download path")
            print('file exist')
            if (os.path.exists(f'{basePath}/Location(s) Holidays.xlsx')):            # check whether the file exist in the "Activity Tracker"
                logger.info(f"file does not exist in the company holiday folder")
                logger.info(f"remove duplicate method called to remove the repeated dates")
                remove_duplicate(f'{basePath}/Location(s) Holidays.xlsx')          # call the remove_current_file(formatted_date) method with the parameter formatted date (MMM DDth YYYY)
            else:
                logger.info(f"move the file from downloads to basepath")
                shutil.move(file_Path, rf'{basePath}/Location(s) Holidays.xlsx')     # move the file from download folder to Activity Tracker folder
                logger.info(f"remove duplicate method called to remove the repeated dates")
                remove_duplicate(f'{basePath}/Location(s) Holidays.xlsx')                                          # if the file does not exist in the downloads folder
            print('file does not exist')
        else:
            webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
            print(f'location holiday file not exist')
            message_content = f'location holiday file not exist'
            # send the respective employee details to the manager to approve the leave request
            data = {
                "content": message_content
            }
            response = requests.post(webhook_url, data=data)
            # Check whether the response send or not
            if response.status_code == 204:
                print("Message sent successfully!")
            else:
                print(f"Failed to send message. Status code: {response.status_code} {response.text}")
    except Exception as e:
        logger.info(f"error occurs in move_activity_file() method {e}")
        webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
        print(f'exception occurs in get_Holiday_List.move_activity_file() \n {e}')
        message_content = f'exception occurs in get_Holiday_List.move_activity_file() \n{e}'
        # send the respective employee details to the manager to approve the leave request
        data = {
            "content": message_content
        }
        response = requests.post(webhook_url, data=data)
        # Check whether the response send or not
        if response.status_code == 204:
            print("Message sent successfully!")
        else:
            print(f"Failed to send message. Status code: {response.status_code} {response.text}")
# End of move_activity_file() method

# Create a main folder with the name "Activity Tracker"
def create_main_folder(current_year):
    try:
        logger.info(f"create_main_folder method gets executed")
        basePath = 'C:/Users/ADMIN/Leave Register'                             # store the base path
        os.mkdir(os.path.join(basePath, f'Company Holiday List {current_year}'))    # create a folder in the base path
        logger.info(f"Company Holiday List {current_year} folder created")
        print('Activity Tracker Folder Created')
        move_activity_file(current_year)                                    # call the move_activity_file() method which is used to move file method to move the file from downloads to "Activity Tracker" folder
    except Exception as e:
        logger.critical(f"error occur in create_main_folder method {e}")
        webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
        print(f'exception occurs in get_Holiday_List.export_excel_file() \n {e}')
        message_content = f'exception occurs in get_Holiday_List.export_excel_file() \n{e}'
        # send the respective employee details to the manager to approve the leave request
        data = {
            "content": message_content
        }
        response = requests.post(webhook_url, data=data)
        # Check whether the response send or not
        if response.status_code == 204:
            print("Message sent successfully!")
        else:
            print(f"Failed to send message. Status code: {response.status_code} {response.text}")
# end of create_main_folder method

def consolidate_Holiday(current_year):
    try:
        logger.info(f"consolidate_Holiday method get executed")
        file_Path = f'C:/Users/ADMIN/Leave Register/Company Holiday List {current_year}'  # store the "Activity Tracker" folder path in the variable
        if (os.path.exists(file_Path)):  # check whether the folder exist or not
            logger.info(f"company holiday list {current_year} folder exist")
            print('file exist')
            move_activity_file(current_year)  # call the move_activity_file() method which is used to move file method to move the file from downloads to "Activity Tracker" folder
        else:  # if the folder does not exist
            logger.info(f"company holiday List {current_year} folder do not exist")
            print('file does not exist')
            logger.info(f"call create_main_folder to create company holiday list folder")
            create_main_folder(current_year)
    except Exception as e:
        logger.critical(f"getting error in consolidate_holiday function {e}")
        webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
        print(f'exception occurs in get_Holiday_List.consoliate_Holiday() \n {e}')
        message_content = f'exception occurs in get_Holiday_List.consoliate_Holiday() \n{e}'
        # send the respective employee details to the manager to approve the leave request
        data = {
            "content": message_content
        }
        response = requests.post(webhook_url, data=data)
        # Check whether the response send or not
        if response.status_code == 204:
            print("Message sent successfully!")
        else:
            print(f"Failed to send message. Status code: {response.status_code} {response.text}")

def export_excel_file(current_year):
    try:
        logger.info("export_excel_file method started execution")
        # Create a new driver instance and add the webdriver
        s = Service('C:/Users/ADMIN/PycharmProjects/leave_register/chromedriver.exe')
        driver = webdriver.Chrome(service=s)

        # Open HRApp
        driver.get('https://digitalseo.hrapp.co/auth')

        # Maximizing the chrome window
        driver.maximize_window()

        sleep(8)
        # click on the sign with email button
        click_EmailBtn = driver.find_element(By.CSS_SELECTOR, '#signinWithEmailBtnText')
        click_EmailBtn.click()

        # Identify and interact with the username field
        enterEmailId = driver.find_element(By.CSS_SELECTOR, '#formSigninEmailId').send_keys(
            'jaya@seofeeds.com')  # Adjust based on actual element ID or another selector
        sleep(1)

        # click on the Next button
        click_NextBtn = driver.find_element(By.CSS_SELECTOR, '#email-verification-button')
        click_NextBtn.click()
        sleep(3)

        # Identify and interact with the password field
        enterUserPassword = driver.find_element(By.ID, 'formSigninPassword').send_keys(
            'Priya446')  # Adjust based on actual element ID or another selector
        sleep(2)

        # Identify and click the signin button
        click_SignInBtn = driver.find_element(By.ID, 'email-password-submit-button')
        click_SignInBtn.click()
        sleep(20)

        logger.info("Signed to HR App to get the holiday list")
        # Identify and click the employee button
        click_EmployeeBtn = driver.find_element(By.CLASS_NAME, 'hr-core-hr')
        click_EmployeeBtn.click()
        sleep(1)

        # inside the employee button we have a option called leave click the click button
        click_EmployeeLeaveBtn = driver.find_element(By.CLASS_NAME, 'hr-core-hr-employee-data-management')
        click_EmployeeLeaveBtn.click()
        sleep(10)

        # inside the employee button we have a option called leave click the click button
        click_Employeeholiday = driver.find_element(By.XPATH, '//*[@id="app"]//div[2]/div/a[3]')
        click_Employeeholiday.click()
        sleep(10)

        # inside the employee button we have a option called leave click the click button
        click_locationTirunelveliBtn = driver.find_element(By.XPATH, '//*[@id="tab-2"]/div[1]/div/div/section/div[1]/div/div[1]/div[3]/div[2]/button/span/i')
        click_locationTirunelveliBtn.click()
        sleep(5)

        # inside the employee button we have a option called leave click the click button
        click_locationBtn = driver.find_element(By.CLASS_NAME, 'fa-file-export')
        click_locationBtn.click()
        sleep(10)
        logger.info("Holiday list downloaded")
        driver.quit()
        consolidate_Holiday(current_year)
    except Exception as e:
        logger.critical(f"exception occurs in the export excel file method {e}")
        webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
        print(f'exception occurs in get_Holiday_List.export_excel_file() \n {e}')
        message_content = f'exception occurs in get_Holiday_List.export_excel_file() \n{e}'
        # send the respective employee details to the manager to approve the leave request
        data = {
            "content": message_content
        }
        response = requests.post(webhook_url, data=data)
        # Check whether the response send or not
        if response.status_code == 204:
            print("Message sent successfully!")
        else:
            print(f"Failed to send message. Status code: {response.status_code} {response.text}")
# export_excel_file(2023)