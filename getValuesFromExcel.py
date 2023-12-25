import requests
import os
from datetime import date, datetime
import shutil
from openpyxl import load_workbook
import logging
import json

# Discord Webhook URL, store today's date,
webhook_url = 'https://discord.com/api/webhooks/1162336441137897613/VfqsEdrd6cXq7nAcLOm5y9z1JusTn1yajQ8rL5MYMoKQeXdC_62Kn0X98dA2wlpfUGy1'
today = str(date.today())
file_Path = f'C:/Users/ADMIN/Leave Register/Leave Register for {today}/Leave Register.xlsx'

def setup_logger(log_file,name):
    logger = logging.getLogger(name)
    # Check if handlers are already added to the logger
    if len(logger.handlers) == 0:
        logger.setLevel(logging.DEBUG)
        handler = logging.FileHandler(log_file)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', '%H:%M')
        handler.setFormatter(formatter)
        logger.addHandler(handler)
    return logger
today = datetime.today()

log_file = f"C:/Users/ADMIN/Log/LeaveRegisterLog/LeaveRegister_getvalues log on {today.date()}.log"
if(os.path.exists(log_file)):
    logger = setup_logger(log_file, "my_logger")
    print(log_file)
else:
    log_file = open(f"C:/Users/ADMIN/Log/LeaveRegisterLog/LeaveRegister_getvalues log on {today.date()}.log", "x")
    logger = setup_logger(log_file.name, "my_logger")
print(logger)

# Send request to discord webhook
def send_discord_message(webhook_url, content):
    data = {
        "content": content
    }
    response = requests.post(webhook_url, data=data)
    return response

def get_column_letter_from_coordinate(coordinate):
    return ''.join(filter(str.isalpha, coordinate))

# Send whatsapp message
def send_whatsapp_message_to_individual(message,empManagerName):
    # API endpoint and token
    logger.info('send whatsapp message to indidvidual method executed')
    message = ', '.join(message)
    logger.info(f'{message}')
    endpoint = "https://graph.facebook.com/v18.0/109611135380129/messages"
    token = "EAAQVfMZBZCG0sBO2g3zmH4UY4fERWFerW7U6XZBMIhjDVKkUdx0tJkNN2d7sLZBZCIyWzT0267Wx2fEHheEZAmcKBzog2DHiDEDfm1WnYUEZCUA7ayl0bVp9ubPR1u8GkbN2CtxSXFZCpy6LREQPxQX3utIxyGxakL2UDkeG3KwdO2XPaAUNP9RsisMO"  # Replace with your token
    file_Path = f'C:/Users/ADMIN/Employee Contact/October/Employee Contact.xlsx'
    employee_contact_book = load_workbook(file_Path)
    employee_contact_sheet = employee_contact_book.active
    logger.info(f'get excel sheet {file_Path}')
    managernumber = ''
    column_letter = get_column_letter_from_coordinate(employee_contact_sheet.cell(row=1, column=2).coordinate)
    for cell in employee_contact_sheet[column_letter]:
        if (employee_contact_sheet.cell(row=cell.row, column=2).value == empManagerName):
            logger.info('get manager number')
            managernumber = employee_contact_sheet.cell(row=cell.row, column=3).value
            print(f'{managernumber} -- {empManagerName}')
            break
    logger.info(f'manager number {managernumber} \n {message}')
    # Message data
    data = {
        "messaging_product": "whatsapp",
        "recipient_type": "individual",
        "to": f"+91 {managernumber}",  # replace f"+91 {managernumber}"
        "type": "template",
        "template": {
            "name": "leave_register",
            "language": {
                "code": "en_US"
            },
            "components": [
                {
                    "type": "body",
                    "parameters": [
                        {"type": "text", "text": f"{empManagerName}"},
                        {"type": "text", "text": f"{message}"}
                    ]
                }
            ]
        }
    }

    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }

    response = requests.post(endpoint, headers=headers, data=json.dumps(data))

    if response.status_code == 200:
        logger.info(f'whatsapp message sent successfully - {message}')
        print("Message sent successfully!")
        print(response.text)
    else:
        logger.error(f'whatsapp message didn\'t send for {managernumber}, {message} \n {response.text}')
        print("Failed to send message:", response.text)

# Send a discord message to the user
def send_discord_message_to_user():
    try:
        logger.info(f'send_discord_message_to_user method gets executed')
        # get the file path URL
        book = load_workbook(file_Path)
        sheet = book.active

        # To store the Leave applied employee details
        empIdList = []
        empNameList = []
        dateList = []
        empManagerList = []
        # Iterate the Applied Satus column
        for cell in sheet[sheet.cell(row=1, column=18).column_letter]:
            # check whether the Applied Status is "Applied"
            if (cell.value == 'Applied'):
                # Get Employee Id who has send the
                logger.info(f'Get the list of employee details who are applied for leave')
                empId = sheet.cell(row=cell.row, column=1).value
                empIdList.append(empId)
                # Get Employee Name who has send the request
                empName = sheet.cell(row=cell.row, column=2).value
                empNameList.append(empName)
                # Get Employee Manager Name who has send the request
                empManager = sheet.cell(row=cell.row, column=5).value
                empManagerList.append(empManager)
                # Get the Date
                date = sheet.cell(row=cell.row, column=10).value
                dateList.append(date)
                logger.info(f'{empId} | {empName} | {empManager} | {date}')
                print(f"Hi {empManager}, \n{empName} has applied for leave on {date}. Please approve the leave.")
        if (len(empId) == 0):
            message = f'Good Morning Team, there are no leave records for approval today.'
            webhook_url = 'https://discord.com/api/webhooks/1162336441137897613/VfqsEdrd6cXq7nAcLOm5y9z1JusTn1yajQ8rL5MYMoKQeXdC_62Kn0X98dA2wlpfUGy1'
            response = send_discord_message(webhook_url, message)
            if response.status_code == 204:
                print("Message sent successfully!")
            else:
                print(f"Failed to send message. Status code: {response.status_code} {response.text}")
            print("\n")
        try:
            # Organizing the data
            organized_data = {}
            for manager, employee, date in zip(empManagerList, empNameList, dateList):
                if manager not in organized_data:
                    organized_data[manager] = {}
                if employee not in organized_data[manager]:
                    organized_data[manager][employee] = set()
                organized_data[manager][employee].add(date)

            # Formatting the message
            messages = []
            managerName = []
            for manager, employees in organized_data.items():
                employee_messages = []
                for employee, dates in employees.items():
                    dates_list = sorted(list(dates))
                    # Check if the employee has multiple dates, and format accordingly
                    if len(dates_list) > 1:
                        last_date = dates_list.pop()
                        dates_str = ', '.join(dates_list) + ' and ' + last_date
                    else:
                        dates_str = dates_list[0]
                    employee_messages.append(f"{employee} has applied for leave on {dates_str}")
                managerName.append(manager)
                print(manager, '--', employee_messages)
                send_whatsapp_message_to_individual(employee_messages, manager)
                messages.append(f"Hi {manager}, \n{', '.join(employee_messages)}. Please approve the leave.")
            empLeaveList = zip(messages,managerName)
            # Print the messages
            for message,manager in empLeaveList:
                print(message)
                # webhook_url = 'https://discord.com/api/webhooks/1162336441137897613/VfqsEdrd6cXq7nAcLOm5y9z1JusTn1yajQ8rL5MYMoKQeXdC_62Kn0X98dA2wlpfUGy1'
                webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
                response = send_discord_message(webhook_url, message)
                if response.status_code == 204:
                    print("Message sent successfully!")
                    # send_whatsapp_message_to_individual(message, manager)
                else:
                    print(f"Failed to send message. Status code: {response.status_code} {response.text}")
                print("\n")
                webhook_url = 'https://discord.com/api/webhooks/1162336441137897613/VfqsEdrd6cXq7nAcLOm5y9z1JusTn1yajQ8rL5MYMoKQeXdC_62Kn0X98dA2wlpfUGy1'
                response = send_discord_message(webhook_url, message)
                if response.status_code == 204:
                    print("Message sent successfully!")
                    # send_whatsapp_message_to_individual(message, manager)
                else:
                    print(f"Failed to send message. Status code: {response.status_code} {response.text}")

        except Exception as e:
            print("exception occurs")
            webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
            logger.info(f'Whatsapp message sent but getting split error')
            if "'list' object has no attribute 'split'" in str(e):
                print("Trying to split a list, which is not supported.")
                result = None
            else:
                logger.critical(f'Error occur when sending whatsapp message')
                print(f"Another AttributeError occurred: {e}")
                print("Something went wrong when writing to the file")
                print(e)
                message_content = f"Message haven't sent in whatsapp group"
                # send the respective employee details to the manager to approve the leave request
                response = send_discord_message(webhook_url, message_content)
                if response.status_code == 204:
                    print("Message sent successfully!")
                else:
                    print(f"Failed to send message. Status code: {response.status_code} {response.text}")
    except Exception as e:
        logger.critical(f'Error occur in the send _discord_message_to_user method')
        webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
        print(f'exception occurs in getValuesFromExcel.send_discord_message_to_user() {e}')
        message_content = f'exception occurs in getValuesFromExcel.send_discord_message_to_user()) {e}'
        # send the respective employee details to the manager to approve the leave request
        data = {
            "content": message_content
        }
        response = requests.post(webhook_url, data=data)
        # Check whether the response send or not
        if response.status_code == 204:
            print("Message sent successfully!")
        else:
            print(f"Failed to send message. Status code: {response.status_code} {response.text}")

# Create a Main Folder Named Leave Register
def create_main_folder():
    try:
        logger.info(f'create_main_folder method get executed')
        basePath = 'C:/Users/ADMIN'
        os.mkdir(os.path.join(basePath, 'Leave Register'))
        logger.info(f'Leave Register folder created')
        print('Folder Created')
        create_sub_folder()
    except Exception as e:
        logger.critical(f'Error occur in the getValuesFromExcel.create_main_folder method {e}')
        webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
        print(f'exception occurs in getValuesFromExcel.create_main_folder() {e}')
        message_content = f'exception occurs in getValuesFromExcel.create_main_folder() {e}'
        # send the respective employee details to the manager to approve the leave request
        data = {
            "content": message_content
        }
        response = requests.post(webhook_url, data=data)
        # Check whether the response send or not
        if response.status_code == 204:
            print("Message sent successfully!")
        else:
            print(f"Failed to send message. Status code: {response.status_code} {response.text}")
# Create a Sub Folder Named Leave Register for today's date and moving the Leave Register file from one location to another
def create_sub_folder():
    try:
        logger.info(f'create_sub_folder method get executed')
        file_Path = 'C:/Users/ADMIN/Downloads/Leave Register.xlsx'
        if (os.path.exists(file_Path)):
            logger.info(f'Leave Register file exist in the download folder')
            print('file exist')
            basePath = 'C:/Users/ADMIN/Leave Register'
            today = str(date.today())
            if (os.path.exists(rf'{basePath}/Leave Register for {today}/Leave Register.xlsx') != True):
                logger.info(f'Leave Register file does not exist the Leave Register for {today} folder')
                if not (os.path.exists(rf'{basePath}/Leave Register for {today}')):
                    os.mkdir(os.path.join(basePath, f'Leave Register for {today}'))
                    logger.info(f'Leave Register for {today} folder created')
                print('Folder Created')
                shutil.move('C:/Users/ADMIN/Downloads/Leave Register.xlsx',
                            rf'{basePath}/Leave Register for {today}/Leave Register.xlsx')
                logger.info(f'File moved from download path to Leave Register folder path ')
            print('File has been moved from source to destination')
            # call the send_discord_message_to_user() function to perform the discord webhook
            send_discord_message_to_user()
        else:
            today = str(date.today())
            print(today)
            if (os.path.exists(rf'C:/Users/ADMIN/Leave Register/Leave Register for {today}/Leave Register.xlsx')):
                send_discord_message_to_user()
            else:
                webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
                print(f'Leave Register file does not exist so execution stopped')
                message_content = f'Leave Register file does not exist so execution stopped'
                # send the respective employee details to the manager to approve the leave request
                data = {
                    "content": message_content
                }
                response = requests.post(webhook_url, data=data)
                # Check whether the response send or not
                if response.status_code == 204:
                    print("Message sent successfully!")
                else:
                    print(f"Failed to send message. Status code: {response.status_code} {response.text}")
    except Exception as e:
        logger.critical(f'Error occur in the getValuesFromExcel.create_sub_folder {e}')
        webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
        print(f'exception occurs in getValuesFromExcel.create_sub_folder() {e}')
        message_content = f'exception occurs in getValuesFromExcel.create_sub_folder() {e}'
        # send the respective employee details to the manager to approve the leave request
        data = {
            "content": message_content
        }
        response = requests.post(webhook_url, data=data)
        # Check whether the response send or not
        if response.status_code == 204:
            print("Message sent successfully!")
        else:
            print(f"Failed to send message. Status code: {response.status_code} {response.text}")
# end of the create_sub_folder() method


# Check whether the file exist or not
def check_folder_Exist_or_not():
    try:
        logger.info(f'check_folder_Exist_or_not method gets executed')
        file_Path = 'C:/Users/ADMIN/Leave Register'
        if (os.path.exists(file_Path)):
            logger.info(f'Leave Register folder exist')
            print('file exist')
            create_sub_folder()
        else:
            logger.info(f'Leave Register folder does not exist')
            print('file does not exist')
            create_main_folder()
    except Exception as e:
        logger.critical(f'Error occurs in the check_folder_exist_or_not method {e}')
        webhook_url = 'https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail'
        print(f'exception occurs in getValuesFromExcel.check_folder_Exist_or_not() {e}')
        message_content = f'exception occurs in getValuesFromExcel.check_folder_Exist_or_not() {e}'
        # send the respective employee details to the manager to approve the leave request
        data = {
            "content": message_content
        }
        response = requests.post(webhook_url, data=data)
        # Check whether the response send or not
        if response.status_code == 204:
            print("Message sent successfully!")
        else:
            print(f"Failed to send message. Status code: {response.status_code} {response.text}")
# end of the check_folder_Exist_or_not() method


